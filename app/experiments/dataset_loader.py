"""
면접 데이터셋 로더
==================
AI_NLP_면접_데이터셋_2단계.xlsx 파일을 로드하여
질문, 평가 기준, 수준별 예시 답변을 제공합니다.
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path

DATASET_PATH = Path(__file__).parent.parent.parent / "AI_NLP_면접_데이터셋_2단계.xlsx"
SHEET_NAME = "AI_NLP_면접_데이터셋"

# 데이터셋 카테고리 → InterviewType.value 매핑
CATEGORY_TO_TYPE: dict[str, str] = {
    "인성": "인성면접",
    "AI/NLP 기초": "기술면접",
    "프로젝트 경험": "기술면접",
    "문제해결/실무": "기술면접",
}

# CompetitorLevel.value → 데이터셋 답변 키 매핑
LEVEL_TO_ANSWER_KEY: dict[str, str] = {
    "high": "high",
    "mid":  "mid",
    "low":  "low",
}


@dataclass
class DatasetQuestion:
    id: int
    category: str
    difficulty: str           # "상" | "중" | "하"
    question: str
    criteria: str
    answers: dict[str, str]   # "high" | "mid" | "low" → 답변 텍스트

    @property
    def interview_type(self) -> str:
        return CATEGORY_TO_TYPE.get(self.category, "기술면접")


class DatasetLoader:
    """Excel 데이터셋을 한 번만 로드하고 캐싱하는 싱글턴 로더."""

    def __init__(self) -> None:
        self._questions: list[DatasetQuestion] = []
        self._loaded = False

    def _load(self) -> None:
        if self._loaded:
            return
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl이 설치되지 않았습니다. `pip install openpyxl`을 실행하세요."
            )

        wb = openpyxl.load_workbook(DATASET_PATH)
        ws = wb[SHEET_NAME]

        for r in range(2, ws.max_row + 1):
            row_id = ws.cell(r, 1).value
            if row_id is None:
                continue
            self._questions.append(
                DatasetQuestion(
                    id=int(row_id),
                    category=ws.cell(r, 2).value or "",
                    difficulty=ws.cell(r, 3).value or "중",
                    question=ws.cell(r, 4).value or "",
                    criteria=ws.cell(r, 5).value or "",
                    answers={
                        "high": ws.cell(r, 6).value or "",
                        "mid":  ws.cell(r, 7).value or "",
                        "low":  ws.cell(r, 8).value or "",
                    },
                )
            )
        self._loaded = True

    def get_all(self) -> list[DatasetQuestion]:
        self._load()
        return list(self._questions)

    def get_by_interview_type(self, interview_type_value: str) -> list[DatasetQuestion]:
        """interview_type_value: '인성면접' | '기술면접' | 'HR면접'"""
        self._load()
        # HR면접은 인성면접 질문을 활용
        target = interview_type_value if interview_type_value != "HR면접" else "인성면접"
        return [q for q in self._questions if q.interview_type == target]

    def get_question_texts(self, interview_type_value: str | None = None) -> list[str]:
        """질문 텍스트 목록 반환 (면접관 에이전트 시드용)."""
        self._load()
        src = (
            self.get_by_interview_type(interview_type_value)
            if interview_type_value
            else self._questions
        )
        return [q.question for q in src]

    def find(self, question_text: str) -> DatasetQuestion | None:
        self._load()
        for q in self._questions:
            if q.question.strip() == question_text.strip():
                return q
        return None

    def get_criteria(self, question_text: str) -> str | None:
        q = self.find(question_text)
        return q.criteria if q else None

    def get_example_qa(
        self,
        interview_type_value: str,
        level: str,
        exclude_question: str | None = None,
    ) -> tuple[str, str] | None:
        """
        해당 면접 유형에서 level에 맞는 예시 (질문, 답변) 쌍을 반환합니다.
        competitor agent 프롬프트의 few-shot 예시로 사용합니다.

        level: 'high' | 'mid' | 'low'
        """
        candidates = self.get_by_interview_type(interview_type_value)
        for q in candidates:
            if exclude_question and q.question == exclude_question:
                continue
            answer = q.answers.get(level, "")
            if answer:
                return (q.question, answer)
        return None


# 싱글턴 인스턴스
dataset_loader = DatasetLoader()
